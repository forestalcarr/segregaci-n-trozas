# Se presenta un modelo de segregación de una troza individual de Pinus radiata para cuatro productos: debobinable, aserrable con poda, aserrable sin poda y pulpable.
# La variación de los productos dependerá principalmente del dap y la altura de poda.
#Se asume un ahusamiento cónico perfecto truncado. Es decir, que se describe por el decrecimiento de una función lineal.

import pandas as pd
import numpy as np
from math import pi
from math import pow 

# variables de entrada
especie = "Pinus radiata"
dap = 60                  # diámetro (cm) a los 1.3 metros
HT = dap/2 + 7.8            # altura total (m)
ahusamiento = 1              # este valor representa el decrecimiento en diámetro en función del aumento en la altura (cm/m)
hpoda = 9                    # altura de poda en (m)
tocón = 0.3                  # pérdida por tocón (m)
pérdida = 0.05              # se refiere a la pérdida de fuste debido al efecto del instrumento de corte al momento del trozado (m)

# Parámetros de los productos
largo_debo = 2.8             # largo de trozas debobinables (m)
largo_ase = 4.1              # largo de ambos tipos de trozas aserrables (m)
largo_pulp = 2.44            # largo de trozas puplables (m)
iu_debo = 32                 # índice de utilización de troza debobinable (cm)
iu_ase = 18                   # índice de utilización para ambos tipos de trozas aserrablese (cm)
iu_pulp = 8                  # índice de utilización de troza pulpable (cm)
largo = tocón


#otros parámetros
dmay = dap - (ahusamiento)
dmen= dmay- (ahusamiento*largo_pulp)


# Definir la estructura del DataFrame

df = pd.DataFrame(columns=['Tipo Troza', 'Dmayor', 'Dmenor','Volumen (m³)'])


def smalian(dmay:float, dmen:float, largo:float):
    " Calcula el volumen de troza utilizando la metodología Smalian "
    vol = round((((dmay** 2 * pi) / 40000 + (dmen**2 * pi) / 40000) / 2) * largo, 3)
    return vol


while largo <= HT and dmen >= iu_pulp:
    if dmay - ahusamiento * largo_debo >= iu_debo and \
    largo + largo_debo <= hpoda and \
    largo + largo_debo <= HT:

        dmen = dmay -(ahusamiento*largo_debo)
        df = df.append({'Tipo Troza': 'Debobinable', 'Dmayor': round(dmay,2), 'Dmenor': round(dmen,2), 'Volumen (m³)': smalian(dmay, dmen, largo_debo)}, ignore_index=True)

        largo = largo + largo_debo
        dmay = (dmen - pérdida)
        
     # ASERRABLE CON PODA
    elif dmay - ahusamiento * largo_ase >= iu_ase and \
        largo + largo_ase <= hpoda and \
        largo + largo_ase <= HT:
                        
        dmen = (dmay - (largo_ase * ahusamiento))

        #indexar valor en la tabla correspondiente
        df = df.append({'Tipo Troza': 'Aserrable CP', 'Dmayor': round(dmay,2), 'Dmenor': round(dmen,2), 'Volumen (m³)': smalian(dmay, dmen, largo_ase)}, ignore_index=True)

        largo = largo + largo_ase
        dmay = (dmen - pérdida) 

     # ASERRABLE SIN PODA
    elif dmay - ahusamiento * largo_ase >= iu_ase and \
        largo + largo_ase <= HT:
        dmen = (dmay - (largo_ase * ahusamiento))

        df = df.append({'Tipo Troza': 'Aserrable SP', 'Dmayor': round(dmay,2), 'Dmenor': round(dmen,2), 'Volumen (m³)': smalian(dmay, dmen, largo_ase)}, ignore_index=True)

        largo = largo + largo_ase
        dmay = (dmen - pérdida)

    # PULPABLE
    elif largo + largo_pulp < HT and \
    dmay - ahusamiento * largo_pulp >= iu_pulp:
                        
        dmen = (dmay - (largo_pulp * ahusamiento))

        df = df.append({'Tipo Troza': 'Pulpable', 'Dmayor': round(dmay, 2), 'Dmenor': round(dmen, 2), 'Volumen (m³)': smalian(dmay, dmen, largo_pulp)}, ignore_index=True)
        largo = largo + largo_pulp
        dmay = (dmen - pérdida)
        #----------# 

    if largo + largo_pulp * ahusamiento >= HT or dmen - largo_pulp * ahusamiento <= iu_pulp:
        break


df['Nro Troza'] = np.arange(1, len(df) + 1)

print(df)

df.to_excel(f'Segregación de troza.xlsx', index=False, sheet_name=f"DAP{dap}")

#! ajustar las celdas al tamaño del texto del DataFrame y centrar los valores
import openpyxl
# Open the workbook and select the worksheet
wb = openpyxl.load_workbook('Segregación de troza.xlsx')
ws = wb[f"DAP{dap}"]


# Iterate over the cells in the worksheet and center and adjust the size of the cells
for row in ws.iter_rows():
    for cell in row:
        cell.font = openpyxl.styles.Font(size=12, bold=True)

# Save the workbook
wb.save('Segregación de troza.xlsx')


# Generar gráfico de barras
import matplotlib.pyplot as plt
import seaborn as sns


paleta_colores = 'colorblind'

#! Genera el gráfico de barras
plt.figure(figsize=(6,5))
plt.title(f'Trozas individuales en árbol de dap:{dap}cm')
sns.barplot(x= 'Nro Troza', y='Volumen (m³)', hue='Tipo Troza', data=df, palette=paleta_colores)

# Abre el archivo Excel en modo escritura
wb = openpyxl.load_workbook(r'Segregación de troza.xlsx')

# Selecciona la pestaña donde quieres guardar el gráfico
ws = wb[f"DAP{dap}"]
# Guarda el gráfico en la pestaña seleccionada
plt.savefig(f'Gráficos/barras_vol_individual dap{dap}.png')
img = openpyxl.drawing.image.Image(f'Gráficos/barras_vol_individual dap{dap}.png')
img.anchor = 'H1'
ws.add_image(img)
# Guarda los cambios en el archivo Excel
wb.save('Segregación de troza.xlsx')

#! Gráfico de tortas con el volmen acumulado según tipo de troza

plt.figure(figsize=(6,5))
plt.title(f'Volumen acumulado (m³) por tipo de troza en árbol de dap:{dap}cm')
# Agrupa el dataframe por el tipo de troza y calcula el volumen acumulado
# Filtra los valores nan del dataframe
df_filtrado = df[df['Volumen (m³)'].notnull()]

# Genera el gráfico de torta usando la columna 'Tipo Troza' como labels y la columna 'Volumen (m³)' como valores
plt.pie(df_filtrado.groupby('Tipo Troza')['Volumen (m³)'].sum(), labels= df_filtrado.groupby('Tipo Troza')['Volumen (m³)'].sum(), autopct='%1.1f%%')


# Abre el archivo Excel en modo escritura
wb = openpyxl.load_workbook(r'Segregación de troza.xlsx')

# Selecciona la pestaña donde quieres guardar el gráfico
ws = wb[f"DAP{dap}"]
# Guarda el gráfico en la pestaña seleccionada
plt.savefig(f'Gráficos/torta_vol_acumulado dap{dap}.png')
img = openpyxl.drawing.image.Image(f'Gráficos/torta_vol_acumulado dap{dap}.png')
img.anchor = 'S1'
ws.add_image(img)
# Guarda los cambios en el archivo Excel
wb.save('Segregación de troza.xlsx')



